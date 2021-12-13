from django.shortcuts import render,reverse,redirect
import pandas as pd
import json
import requests
import openpyxl
from datetime import date
from django.contrib import messages
import uuid
from django.shortcuts import get_object_or_404
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth import authenticate,login
from django.contrib.auth.models import User
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from copon.models import copons,profile,Payment_details,Payment_details_customer,Subscription_details
from django.http import JsonResponse
from django.conf import settings
import stripe
from django.urls import reverse
from django.views.generic import TemplateView
stripe.api_key = get_env_variable('STRIPE_SECRET_KEY')
from datetime import datetime,timedelta
from django.contrib.auth.models import User 
from django.contrib.auth import authenticate, login as django_login,logout
from django.contrib.auth.decorators import login_required
from .models import *


# Create your views here.


def whatsapp(request):
  #requesring  coupon url
  camp=request.GET.get('coupen_url')
  #requesting pretext
  pretext=request.GET.get('pretext')
  return render(request,'whatsapp.html', {'camp_url': camp})


def whatsappdone(request):

  
  url = "https://waba-sandbox.360dialog.io/v1/messages"
  number=request.GET.get('number')
  cam_url=request.GET.get('cam_url')
  pretext=request.GET.get('pretext')
  print(pretext)
  print("7272727272727277")
  posttext=request.GET.get('posttext')
  payload = {
        "to": number,
       "type": "text",
        "text": {
            "body": pretext + "\n"+cam_url + "\n"+posttext,
             
               }
          }
  headers = {
    'D360-Api-Key': "1zaWyR_sandbox",
    'Content-Type': "application/json",
  }

  response = requests.request("POST", url, headers=headers, data= json.dumps(payload))
  return render(request,'whatsappdone.html') 


 
def home(request):
  
  tags=request.GET.get('tags')
  url="https://api4coupons.com/v3/coupon/list"
  payload = json.dumps({
    "client_id": "6384889747946715134741991478448",
    "client_secret": "Pnhughk8ZhTJGxQkyhtc95TUVgPMtuE",
    "tags": tags
  })
  headers = {
  'Content-Type': 'application/json',
  'Cookie': 'PHPSESSID=8a3edbc2d235fdc4c16511cb4afa766d'
  }

  response = requests.request("POST", url, headers=headers, data=payload)
  data=response.json()
  
  
  coupon_info=data['coupon_info']


  context={
  'coupon_info': coupon_info
    }
  if request.user.is_authenticated:
        prof = profile.objects.filter(user=request.user).first()
        request.session['prof'] = prof.pro
        print(prof)
        print("0000000000000000000000000000000000000")
       
        
  
  return render(request,'home.html',context)
#def sendsms(request):
#  camp=request.GET.get('coupen_id')
#  return render(request,'sendsms.html', {'camp_id': camp})

def smsdatabase(request):
  
  camp=request.GET.get('coupen_id')
  phonenumberlist = copons.objects.all()
  return render(request,'smsdatabase.html', {'camp_id': camp, 'copons':  phonenumberlist })
def smsdatabasedone(request):
 phonenumberlist = copons.objects.all()
 cam_id=request.GET.get('coupen_id')
 url = "https://api4coupons.com/v3/send/sms"
 for j in phonenumberlist:
      payload = json.dumps({
      "client_id": "6384889747946715134741991478448",
        "client_secret": "Pnhughk8ZhTJGxQkyhtc95TUVgPMtuE",
         "campaign": cam_id,
         "phone": j.cnumber,
       "sender": "APPOTP"
       })
      headers = {
      'Content-Type': 'application/json',
        'Cookie': 'PHPSESSID=6b8d40e13338ebdda12d3da572707bba'
         }

      response = requests.request("POST", url, headers=headers, data=payload)
      return render(request,'smsdatabasedone.html')
@csrf_exempt 

#def smsdone(request):
#  excel_file = request.FILES["file"]
#  cam_id=request.POST.get('cam_id')
#  url = "https://api4coupons.com/v3/send/sms"

#  wb = openpyxl.load_workbook(excel_file)
#  worksheet = wb["Sheet1"]
#  for row in worksheet.iter_rows():
#    for cell in row:
#      payload = json.dumps({
#      payload = json.dumps({
#      "client_id": "6384889747946715134741991478448",
#      "client_secret": "Pnhughk8ZhTJGxQkyhtc95TUVgPMtuE",
#      "campaign": cam_id,
#      "phone": cell.value,
#      "sender": "APPOTP"
#      })
#      headers = {
 #     'Content-Type': 'application/json',
 #     'Cookie': 'PHPSESSID=6b8d40e13338ebdda12d3da572707bba'
 #     }

 #     response = requests.request("POST", url, headers=headers, data=payload)
 #     return render(request,'smsdone.html')  
def sendemail(request):
    camp=request.GET.get('coupen_id')
    return render(request,'sendemail.html', {'camp_id': camp})

      
@csrf_exempt 

def emaildone(request):
    
    excel_file = request.FILES["file"]
    cam_id=request.GET.get('cam_id')
    url = "https://api4coupons.com/v3/send/email"

    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
      for cell in row:
        payload = json.dumps({
        "client_id": "6384889747946715134741991478448",
        "client_secret": "Pnhughk8ZhTJGxQkyhtc95TUVgPMtuE",
        "campaign": cam_id,
        "email": cell.value
        })
        headers = {
        'Content-Type': 'application/json',
        'Cookie': 'PHPSESSID=6b8d40e13338ebdda12d3da572707bba'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        return render(request,'emaildone.html')  



#for payments



def mainsuccess(request):
 
  return render(request,'mainsuccess.html')



def landingcopon(request):
    
    if request.method == 'POST':
      login_data = request.POST.dict()
      print( login_data)
      p=request.POST.get('address')
      print(p)
      print("555555555555555555555555555555555555555555555555555555")
      pi=request.POST.get('paymentstatus')
      print(pi)
      
      print(list(request.POST.items()))
      
     
      print("[[[[[[[[[[[]]]]]]]]]]]]]]")
     # data = json.loads(request.body)
     # print(data)
      print("$$44444444444444444")
     # Create a PaymentIntent with the order amount and currency
      req_json = json.loads(request.body)
      
      
      print("0000000000000000")

      
      customer = stripe.Customer.create(email=req_json['email'],
      
      name=req_json['name'],)
      print(customer)
    
      intent = stripe.PaymentIntent.create(

           customer=customer,
           amount='2000',
            
           currency='AED',
           description="order for",

           payment_method_types=[

                 'card',
                ],
              
       )
      
      print(req_json['name'])
      print(req_json['email'])
      print("id:",intent['id'])
      print("id:",intent['id'])
      saverecord=Payment_details()
      saverecord.customer_name=req_json['name']
      saverecord.customer_email=req_json['email']
      saverecord.payment_id=intent['id']
      saverecord.save()
      
     
     
      return JsonResponse({
         

         'client_secret' : intent['client_secret'], 'id' : intent['id'],

      })
     
    
    return render(request,'landingcopon.html',)


    '''if request.POST.get('cardname') and request.POST.get('email') and request.POST.get('address') and request.POST.get('city') and request.POST.get('country') and request.POST.get('phonenumber') and request.POST.get('state') and request.POST.get('postcode'):
         saverecord=Payment_details_customer()
        
         saverecord.cardname=request.POST.get('cardname')
         
         
         saverecord.email=request.POST.get('email')
         saverecord.address=request.POST.get('address')
         saverecord.city=request.POST.get('city')
         saverecord.country=request.POST.get('country') 
         saverecord.phonenumber=request.POST.get('phonenumber') 
         saverecord.state=request.POST.get('state') 
         saverecord.postcode=request.POST.get('postcode') 
         saverecord.save()'''


def successlanding(request):

   
   
    
    if request.method == 'POST':
      #Payment=self.get_object
      
      
  
      
      d=json.dumps(request.POST)
      pythonObj = json.loads(d)
      name = pythonObj['phonenumber']
      print(name)
      print("222222222222222222222222222222222222222222222222222222222222222")
      print(d)
      gg=d[5]
      print("444444444444444444444444444444444444444444444444444444444444444444444444444")
      print(gg)
      
      print("999999999999999999999")
     
     


      login_data = request.POST.dict()
      print( login_data)

     
      print(list(request.POST.items()))
      

      print("777777777777777777777777")
      
      p=request.POST.get('address')
      print(p)
      print("555555555555555555555555555555555555555555555555555555")
      pi=request.POST.get('paymentstatus')
      print(pi)


      saverecord=Payment_details_customer()
        
      saverecord.cardname=request.POST.get('cardname')

      u_id= uuid.uuid4()
      print(u_id)
         
      
      saverecord.email=request.POST.get('email')
      saverecord.address=request.POST.get('address')
      saverecord.city=request.POST.get('city')
      saverecord.country=request.POST.get('country') 
      saverecord.phonenumber=request.POST.get('phonenumber') 
      saverecord.state=request.POST.get('state') 
      saverecord.postcode=request.POST.get('postcode') 
      saverecord.paymentstatus=request.POST.get('paymentstatus') 
      saverecord.paymentid=request.POST.get('payment') 
      saverecord.save()
      
      
      cardname=request.POST.get('cardname')
      today = date.today()
      d1 = today.strftime("%d/%m/%Y")
      pay=request.POST.get('payment')
      
     

    return render(request,'successlanding.html',{'name': cardname,'dat':d1,'paym':pay,'Payment':Payment}) 
def okay(request,id):
  
  Payment=Payment_details_customer.objects.get(id=id)

  
  return render(request,'okay.html',{'Payment':Payment})   
    
def success(request):
   
  
  
  
  
  print("222@22222")  
  amount=request.GET.get('amount')
  print(amount)
  print("55555555555")
  session_id=request.GET.get('session_id')
  print(session_id)
  current_user = request.user
  # profile=profile.objects.filter(user=current_user).first()
  #  print("212111111111111111111111")
  prof = profile.objects.filter(user=request.user).first()
  if amount == "100000":
    
    prof.pro = True
    prof.subscriptiontype="M"
    prof.save()
    return redirect('home')
    #return render(request,'success.html') 
  elif amount == "1200000":
      
      prof.pro = True
      prof.subscriptiontype="Y"
      
      prof.save()   
      return redirect('home')
     # return render(request,'home.html') '''
 
  
  #return render(request,'success.html')   



      
def cancel(request):
    pass
    #return render(request,'cancel.html')  
'''def success(request):    
  
  
  membership = request.POST.get('membership')
  if membership == "MONTHLY":
    price='price_1JjibCIhoVgdSGghU4wV8EMC'
    amount=1000

  else:
    price ='price_1JjibCIhoVgdSGghU4wV8EMC'
    amount=12000

  if request.method == 'POST':  
   print("99999999999999999999")
   req_json = json.loads(request.body)
   customer = stripe.Customer.create(email=req_json['email'],
      
   name=req_json['name'],)
  
   print("222222222222222")
   print(customer["id"])

   
   print("4444444444")
    
   intent = stripe.Subscription.create(

           customer=customer["id"],
           payment_behavior='default_incomplete',
           items=[
    {"price": price},
        ],
       )
  return JsonResponse(intent) 
  
  return render(request,'success.html',)'''

def final(request): 
   return render(request,'final.html',)


def checkout(request):
  
 

      
  
  
  membership = request.POST.get('membership')
  

  if membership == "MONTHLY":
    price='price_1JjibCIhoVgdSGghU4wV8EMC'
    amount=1000

  else:
    price ='price_1JjibCIhoVgdSGghU4wV8EMC'
    amount=12000
    
  
    
   
  
  customer=stripe.Customer.create(
  name=request.user,
   )
  checkout_session = stripe.checkout.Session.create(
            #+"/" + "session_id={CHECKOUT_SESSION_ID}")
            success_url=("http://127.0.0.1:8000/success?amount=" + str(amount*100)),
            cancel_url='http://127.0.0.1:8000/cancel',
            payment_method_types=['card'],
            customer=customer,
            
            mode='subscription',
            # automatic_tax={'enabled': True},
            line_items=[{
                'price': price,
                'quantity': 1
            }],
        )
  ema=request.user.email
  print(ema)

  ai=checkout_session['id']
  print(ai)
  print("4444444444444")
  pi=str(request.user)
  ci=checkout_session['customer']
  print(ci)
  print("333333333333333")
  ei=request.user.email
  sava=Subscription_details()
  sava.sname=pi
  sava.email=ei
  sava.customer_id=ci
  sava.session_id=ai
  sava.subscription_status="success"
  
  sava.save()

  print(282828282828)  
      

  return redirect(checkout_session.url, code=303)   

  
   
  
'''def checkout(request):
  membership = request.POST.get('membership')
  

  if membership == "MONTHLY":
    amount=1000
  else:
    amount = 12000
  session = stripe.checkout.Session.create(
    payment_method_types=['card'],
    line_items=[{
    'price_data': {
      'currency': 'usd',
        'product_data': {
        'name': 'coupons',
        },
      'unit_amount': amount*100,
  },
    'quantity': 1,
    }],
    mode='payment',
    success_url=("http://127.0.0.1:8000/success?amount=" + str(amount*100)),
    cancel_url='http://127.0.0.1:8000/failed',  
    )
  return redirect(session.url, code=303)  '''


def subscribe(request):   
 
  
   
  return render(request,'subscribe.html') 
     


def subcompleted(request):

  return render(request,'subcompleted.html') 

       
def logout_account(request):
  logout(request)
  return redirect('signup')

def unsubscribe(request):
  print("44444444444444")
 
  a= Subscription_details.objects.all()
  for j in a:
    if j.email==request.user.email:
      print(j.email)
      print("3333333333333333333")
      print(j.customer_id)
      customer=j.customer_id
      print("4444444444444")
      print(j.subscription_status)
      
     
  session = stripe.billing_portal.Session.create(
      customer= customer,
    return_url='http://127.0.0.1:8000/updated',
  )
  return redirect(session.url)

def updated(request):
  
    prof = profile.objects.filter(user=request.user).first()
  
   
    prof.pro = False
    print(prof.pro)
    print("33333333333")
    prof.subscriptiontype=" "
    prof.save()
    a= Subscription_details.objects.all()
    
    for j in a:
     if j.email==request.user.email:

       print(j.email)
       print("3333333333333333333")
       print(j.customer_id)
       customer=j.customer_id
       print("4444444444444")
       print(j.subscription_status)
      
       b=Subscription_details.objects.filter(email=request.user.email).update(subscription_status="failure")
       
     
       print("ggggggggggggggggggggg")
       print(j.subscription_status)
       


       return redirect('index') 
       
      
    



  
  
  
  
''' print("44444444444444")
  aha= Subscription_details.objects.all()
  print(aha)

  customer= Subscription_details.objects.filter(customer_id=request.user.customer_id)
  print(customer)
  email = request.user.email
  if email:
    customerid=Subscription_details.objects.get(customer_id=request.user.customer_id)

  em=Subscription_details.objects.get(email=request.user.email)
  print(customer)
  


 print("77777777777777")
  session = stripe.billing_portal.Session.create(
    customer= customer,
    return_url='https://example.com/account',
  )
  return redirect(session.url)'''
def signup(request):
  if request.method == 'POST':
  

    print("[[[[[[[[[[[[t")
    username= request.POST.get('input_text')
    print(username)
    print("44444444444")
    password= request.POST.get('input_password')
    print(password)
   
    user=authenticate(username=username,password=password)
    print(user)
    
    if user is not None:
        login(request,user)  
        return redirect('index')
    else:
      messages.warning(request, 'form is not valid.') 
  context={}    
  return render(request,'signup.html')   

  
@csrf_exempt    
def index(request):
   if  request.user.is_authenticated  :
     print("uuuuuuuuuuuuuu")
   else:
     print("777777777777")  
   if request.user.is_authenticated:
        prof = profile.objects.filter(user=request.user).first()
        request.session['prof'] = prof.pro
        print(prof)
        print("0000000000000000000000000000000000000")  
   return render(request,'index.html')
@csrf_exempt    
def reg(request):
   if request.method=='POST':
    print("[[[[[[[[[[[[[[[[[[[[[")
    
    print("222222222")
    #first=request.POST['name']
    first=request.POST['name']
    print(first)
    print("&???????????????????????????????????????????????????")
    
    username=request.POST['username']
    print(username)
    email= request.POST['email']
    print(email)
    password= request.POST['password']
    print(password)
   
    user= User.objects.filter(username=username).first()
    print(user)
    #user=User.objects.get(username=username)
    #user=User.objects.filter(email=email)
    print("----------------")
    
    print("000000000000000")
    
    if User.objects.filter(username=username).exists():
      context = {'message' : 'User already exists' , 'class' : 'danger'}
      return render(request,'reg.html' , context)
    elif User.objects.filter(email=email).exists():
      context = {'message' : 'User already exists' , 'class' : 'danger'}
      return render(request,'reg.html' , context)

    else :
      context = {'message' : 'User created successfully' , 'class' : 'success'}
      user = User.objects.create_user(username = username, email= email,first_name=first,password=password,)
      user.set_password(password)
      user.save()
      
     
      prof = profile(user=user) 
      prof.save()
      return redirect('signup')
     # return render(request,'cancel.html' , context)
        
 


   return render(request,'reg.html')






#  




